"""数据导出服务"""
import csv
import json
from datetime import datetime
from typing import List, Dict, Any
from io import StringIO, BytesIO
from sqlmodel import Session, select

from app.models import Leave, Student, Reviewer, Teacher, Course


class ExportService:
    """数据导出管理服务"""

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str) -> BytesIO:
        """
        导出数据为CSV格式

        Args:
            data: 要导出的数据列表
            filename: 导出文件名

        Returns:
            BytesIO对象，包含CSV数据
        """
        if not data:
            raise ValueError("没有数据可导出")

        output = BytesIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        output.seek(0)

        return output

    @staticmethod
    def export_to_json(data: List[Dict[str, Any]], filename: str) -> BytesIO:
        """
        导出数据为JSON格式

        Args:
            data: 要导出的数据列表
            filename: 导出文件名

        Returns:
            BytesIO对象，包含JSON数据
        """
        if not data:
            raise ValueError("没有数据可导出")

        output = BytesIO()
        json_data = json.dumps(data, ensure_ascii=False, indent=2)
        output.write(json_data.encode('utf-8'))
        output.seek(0)

        return output

    @staticmethod
    def export_leaves_to_excel(
        leaves: List[Leave],
        session: Session
    ) -> BytesIO:
        """
        导出请假数据为Excel格式

        Args:
            leaves: 请假记录列表
            session: 数据库会话

        Returns:
            BytesIO对象，包含Excel数据
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "请假记录"

        # 定义表头
        headers = [
            "请假编号", "学生姓名", "请假日期", "请假天数", "请假类型",
            "状态", "审核员", "审核时间", "备注"
        ]

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # 获取关联数据
        student_ids = {leave.student_id for leave in leaves if leave.student_id}
        reviewer_ids = {leave.reviewer_id for leave in leaves if leave.reviewer_id}

        students = {}
        if student_ids:
            student_results = session.exec(
                select(Student).where(Student.student_id.in_(student_ids))
            ).all()
            students = {s.student_id: s.student_name for s in student_results}

        reviewers = {}
        if reviewer_ids:
            reviewer_results = session.exec(
                select(Reviewer).where(Reviewer.reviewer_id.in_(reviewer_ids))
            ).all()
            reviewers = {r.reviewer_id: r.reviewer_name for r in reviewer_results}

        # 写入数据
        for row_num, leave in enumerate(leaves, 2):
            ws.cell(row=row_num, column=1, value=leave.leave_id)
            ws.cell(row=row_num, column=2, value=students.get(leave.student_id, "未知"))
            ws.cell(row=row_num, column=3, value=leave.leave_date.strftime("%Y-%m-%d %H:%M:%S") if leave.leave_date else "")
            ws.cell(row=row_num, column=4, value=leave.leave_days)
            ws.cell(row=row_num, column=5, value=leave.leave_type or "")
            ws.cell(row=row_num, column=6, value=leave.status)
            ws.cell(row=row_num, column=7, value=reviewers.get(leave.reviewer_id, ""))
            ws.cell(row=row_num, column=8, value=leave.audit_time.strftime("%Y-%m-%d %H:%M:%S") if leave.audit_time else "")
            ws.cell(row=row_num, column=9, value=leave.remarks or "")

        # 自动调整列宽
        for column in ws.columns:
            max_length = max(
                len(str(cell.value)) for cell in column if cell.value
            )
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def generate_filename(prefix: str, format: str) -> str:
        """生成导出文件名"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{format}"